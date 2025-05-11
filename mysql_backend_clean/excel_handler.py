from typing import Dict, List, Any, Optional, Tuple
import base64
import pandas as pd
import logging
import io
import re
import zipfile
import tempfile
import os

try:
    from .db_connection import DBConnection
except ImportError:
    from db_connection import DBConnection

logger = logging.getLogger(__name__)

class ExcelHandler:
    """Excel 檔案處理類"""
    @staticmethod
    def save_questions_to_db(questions: List[Dict[str, Any]], file_id: str, file_name: str, doc_type: str) -> bool:
        """將題目保存到資料庫
        
        Args:
            questions: 題目列表
            file_id: 檔案 ID
            file_name: 檔案名稱
            doc_type: 文件類型
            
        Returns:
            bool: 是否成功保存
        """
        try:
            db = DBConnection()
            # 插入題目到資料庫
            result = db.insert_questions(questions, file_name, doc_type)
            
            if result > 0:
                # 記錄檔案上傳信息到 uploaded_files 表
                upload_result = db.record_file_upload(file_id, file_name, doc_type, result)
                if upload_result:
                    logger.info(f"成功記錄檔案上傳信息：ID={file_id}, 名稱={file_name}, 類型={doc_type}, 題目數={result}")
                else:
                    logger.warning(f"記錄檔案上傳信息失敗：ID={file_id}, 名稱={file_name}")
            
            db.close()
            return result > 0
        except Exception as e:
            logger.error(f"保存題目到資料庫時發生錯誤: {str(e)}")
            return False
    
    @staticmethod
    def get_file_list(doc_type: str = None) -> List[Dict[str, Any]]:
        """獲取檔案列表
        
        Args:
            doc_type: 文件類型過濾條件，如果為 None 則獲取所有類型
            
        Returns:
            List[Dict[str, Any]]: 檔案列表
        """
        try:
            db = DBConnection()
            file_list = db.get_file_list(doc_type)
            db.close()
            return file_list
        except Exception as e:
            logger.error(f"獲取檔案列表時發生錯誤: {str(e)}")
            return []
    
    @staticmethod
    def get_questions(file_id: str) -> Tuple[List[Dict[str, Any]], str]:
        """獲取指定檔案 ID 的題目
        
        Args:
            file_id: 檔案 ID
            
        Returns:
            Tuple[List[Dict[str, Any]], str]: (題目列表, 檔案名稱)
        """
        try:
            db = DBConnection()
            questions, file_name = db.get_questions(file_id)
            db.close()
            return questions, file_name
        except Exception as e:
            logger.error(f"獲取題目時發生錯誤: {str(e)}")
            return [], ""
    
    @staticmethod
    def delete_file(file_id: str) -> bool:
        """刪除檔案及其題目
        
        Args:
            file_id: 檔案 ID
            
        Returns:
            bool: 是否成功刪除
        """
        try:
            db = DBConnection()
            result = db.delete_file_and_questions(file_id)
            db.close()
            return result
        except Exception as e:
            logger.error(f"刪除檔案時發生錯誤: {str(e)}")
            return False
    
    @staticmethod
    def update_questions(file_id: str, questions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """更新題目內容
        
        Args:
            file_id: 檔案 ID
            questions: 要更新的題目列表
            
        Returns:
            Dict: 包含更新結果的字典
        """
        try:
            db = DBConnection()
            result = db.update_questions(file_id, questions)
            db.close()
            return result
        except Exception as e:
            logger.error(f"更新題目時發生錯誤: {str(e)}")
            return {
                'success': False,
                'updated_count': 0, 
                'deleted_count': 0,
                'error': str(e)
            }
    
    @staticmethod
    def extract_images_from_excel_zip(file_path_or_buffer):
        """Extract images directly from Excel file by treating it as a ZIP archive."""
        images = []
        
        try:
            # If it's a file-like object, write it to a temporary file first
            if hasattr(file_path_or_buffer, 'read'):
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file_path = temp_file.name
                temp_file.close()
                
                with open(temp_file_path, 'wb') as f:
                    f.write(file_path_or_buffer.read())
                
                file_path = temp_file_path
            else:
                file_path = file_path_or_buffer
                
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # Find media files
                files = zip_ref.namelist()
                media_files = [f for f in files if f.startswith('xl/media/')]
                logger.info(f"Found {len(media_files)} media files in Excel ZIP structure")
                
                for media_file in media_files:
                    try:
                        # Extract media file content
                        image_data = zip_ref.read(media_file)
                        image_name = os.path.basename(media_file)
                        
                        images.append({
                            'name': image_name,
                            'data': image_data,
                            'size': len(image_data)
                        })
                        
                        logger.info(f"Extracted image: {image_name} ({len(image_data)} bytes)")
                    except Exception as e:
                        logger.error(f"Error extracting {media_file}: {str(e)}")
            
            # Clean up temporary file if created
            if hasattr(file_path_or_buffer, 'read') and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
                        
        except Exception as e:
            logger.error(f"Error opening Excel as ZIP: {str(e)}")
        
        return images

    @staticmethod
    def extract_images_with_openpyxl(file_path_or_buffer):
        """Extract images from Excel file using openpyxl library."""
        images = []
        
        try:
            # If it's a file-like object, write it to a temporary file first
            if hasattr(file_path_or_buffer, 'read'):
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file_path = temp_file.name
                temp_file.close()
                
                with open(temp_file_path, 'wb') as f:
                    f.write(file_path_or_buffer.read())
                
                file_path = temp_file_path
            else:
                file_path = file_path_or_buffer
            
            # Load workbook
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(file_path)
                logger.info(f"Loaded workbook with {len(workbook.sheetnames)} sheets")
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    logger.info(f"Processing sheet: {sheet_name}")
                    
                    # Access images in the sheet
                    if hasattr(sheet, '_images'):
                        for image in sheet._images:
                            try:
                                img_name = f"img_{len(images) + 1}"
                                if hasattr(image, 'name') and image.name:
                                    img_name = image.name
                                
                                # Get image data
                                img_data = None
                                
                                # Try different ways to access the image data
                                if hasattr(image, '_data'):
                                    img_data = image._data()
                                    logger.info(f"Got image data from _data method: {len(img_data)} bytes")
                                elif hasattr(image, 'data'):
                                    img_data = image.data
                                    logger.info(f"Got image data from data attribute: {len(img_data)} bytes")
                                
                                if img_data:
                                    images.append({
                                        'name': img_name,
                                        'data': img_data,
                                        'size': len(img_data),
                                        'sheet': sheet_name
                                    })
                            except Exception as e:
                                logger.error(f"Error processing image in {sheet_name}: {str(e)}")
            except ImportError:
                logger.warning("openpyxl library not available, skipping this method")
            
            # Clean up temporary file if created
            if hasattr(file_path_or_buffer, 'read') and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            
        except Exception as e:
            logger.error(f"Error using openpyxl: {str(e)}")
        
        return images
    
    @staticmethod
    def parse_excel_to_questions(file_content: str, file_type: str) -> List[Dict[str, Any]]:
        """解析 Excel 檔案內容為題目列表
        
        Args:
            file_content (str): Base64 編碼的檔案內容
            file_type (str): 檔案類型
            
        Returns:
            List[Dict[str, Any]]: 題目列表
        """
        try:
            # 解碼 Base64 內容
            content = file_content
            if "base64," in content:
                content = content.split("base64,")[1]
            
            binary_data = base64.b64decode(content)
            buffer = io.BytesIO(binary_data)
            
            # 使用 pandas 讀取 Excel
            df = pd.read_excel(buffer)
            
            # 重置 buffer 以便後續提取圖片
            buffer.seek(0)
            
            # 提取 Excel 中的圖片
            # 首先嘗試使用 openpyxl 提取
            images_from_openpyxl = ExcelHandler.extract_images_with_openpyxl(buffer)
            
            # 重置 buffer 以便使用 ZIP 方法提取
            buffer.seek(0)
            
            # 然後嘗試使用 ZIP 方法提取
            images_from_zip = ExcelHandler.extract_images_from_excel_zip(buffer)
            
            # 合併所有圖片結果 (優先使用 openpyxl 圖片)
            all_images = images_from_openpyxl + images_from_zip
            logger.info(f"從 Excel 檔案中提取了 {len(all_images)} 張圖片")
            
            # 檢查必要欄位是否存在
            required_fields = [
                "QuestionNo.", "ChapterNo", "QuestionInChinese", 
                "AnswerAInChinese", "AnswerBInChinese", "AnswerCInChinese", "AnswerDInChinese", 
                "CorrectAnswer", "AnswerExplainInChinese"
            ]
            
            # 添加可選欄位
            optional_fields = ["Picture", "PIC"]
            
            # 調整欄位名稱，處理可能的空格或大小寫差異
            df.columns = [col.strip() for col in df.columns]
            column_mapping = {}
            
            # 輸出檢查
            logger.info(f"Excel 檔案的欄位名稱: {list(df.columns)}")
            
            for col in df.columns:
                # 處理必要欄位
                for req in required_fields:
                    if col.lower() == req.lower() or col.lower().replace(" ", "") == req.lower().replace(".", ""):
                        column_mapping[col] = req
                
                # 特別處理 PIC 欄位映射到 Picture
                if col.lower() == "pic":
                    column_mapping[col] = "Picture"
                elif col.lower() == "picture":
                    column_mapping[col] = "Picture"
            
            # 檢查是否所有必要欄位都存在 (除了Picture欄位可選外)
            if len(column_mapping) < len(required_fields):
                missing_fields = [field for field in required_fields if field not in column_mapping.values()]
                logger.error(f"Excel 檔案欄位不符合要求，缺少必要欄位: {missing_fields}，目前找到 {list(df.columns)}")
                raise ValueError(f"Excel 檔案格式不正確，請確保包含所有必要欄位")
                
            # 重命名欄位以符合標準格式
            df = df.rename(columns=column_mapping)
            
            # 檢查是否有 "Picture" 欄位
            has_picture = "Picture" in df.columns
            
            # 統計總行數、有效行數和跳過行數
            total_rows = len(df)
            valid_rows = 0
            skipped_rows = 0
            skipped_reasons = {}  # 用於記錄跳過的原因
            
            logger.info(f"開始處理 Excel 檔案，總計 {total_rows} 行數據")
            
            questions = []
            for idx, row in df.iterrows():
                # 檢查題目欄位 - 這是唯一必須的欄位
                if "QuestionInChinese" not in df.columns or pd.isna(row.get("QuestionInChinese")):
                    reason = "缺少題目內容"
                    skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1
                    logger.warning(f"第 {idx+1} 行跳過: {reason}")
                    skipped_rows += 1
                    continue
                
                # 其他欄位都採用寬鬆處理，使用預設值填充而非跳過
                warning_fields = []
                for field in required_fields:
                    if field != "QuestionInChinese" and (field not in df.columns or pd.isna(row.get(field))):
                        warning_fields.append(field)
                
                # 只記錄警告，但仍處理該題目
                if warning_fields:
                    logger.warning(f"第 {idx+1} 行警告: 欄位 {', '.join(warning_fields)} 為空，使用預設值處理")
                
                # 識別正確答案
                correct_answer = row["CorrectAnswer"]
                if isinstance(correct_answer, str) and correct_answer.upper() in ["A", "B", "C", "D"]:
                    letter_to_option = {
                        "A": "AnswerAInChinese",
                        "B": "AnswerBInChinese",
                        "C": "AnswerCInChinese",
                        "D": "AnswerDInChinese"
                    }
                    correct_answer_text = row[letter_to_option[correct_answer.upper()]]
                else:
                    correct_answer_text = correct_answer
                
                # 安全獲取欄位值，對於缺失的欄位使用預設值
                def safe_get(field_name, default=""):
                    if field_name in df.columns and not pd.isna(row.get(field_name)):
                        return str(row.get(field_name))
                    return default
                
                # 構建題目資料
                question = {
                    "question_no": safe_get("QuestionNo.", str(idx+1)),
                    "chapter_no": safe_get("ChapterNo", "未分類"),
                    "question_text": str(row["QuestionInChinese"]),
                    "option_a": safe_get("AnswerAInChinese"),
                    "option_b": safe_get("AnswerBInChinese"),
                    "option_c": safe_get("AnswerCInChinese"),
                    "option_d": safe_get("AnswerDInChinese"),
                    "correct_answer": str(correct_answer_text),
                    "explanation": safe_get("AnswerExplainInChinese", "無解釋"),
                    "picture": None
                }
                
                # 從提取的圖片中分配圖片（如果有）
                if idx < len(all_images):
                    question["picture"] = all_images[idx]["data"]
                    logger.info(f"題目 {question['question_no']} 已分配圖片，大小: {len(question['picture'])} 位元組")
                else:
                    # 如果沒有從 Excel 結構中提取到圖片，則檢查單元格中的圖片欄位
                    picture_field = None
                    for col in df.columns:
                        if col.lower() == "picture" or col.lower() == "pic":
                            picture_field = col
                            break
                    
                    if picture_field and not pd.isna(row.get(picture_field)):
                        # 處理儲存格中的圖片資料
                        pic_data = row[picture_field]
                        logger.info(f"圖片資料類型: {type(pic_data)}")
                        
                        if isinstance(pic_data, str):
                            # 如果是 Base64 格式
                            if pic_data.startswith("data:image") and ";base64," in pic_data:
                                try:
                                    base64_content = pic_data.split(";base64,")[1]
                                    question["picture"] = base64.b64decode(base64_content)
                                    logger.info(f"成功從 Base64 提取圖片: {len(question['picture'])} 位元組")
                                except Exception as e:
                                    question["picture"] = None
                                    logger.warning(f"無法解碼 Base64 圖片: {str(e)}")
                            else:
                                logger.info(f"字串圖片內容片段: {pic_data[:100] if len(pic_data) > 100 else pic_data}")
                                try:
                                    question["picture"] = pic_data.encode("utf-8")
                                    logger.info(f"將字符串轉換為二進位: {len(question['picture'])} 位元組")
                                except Exception as e:
                                    question["picture"] = None
                                    logger.warning(f"無法編碼圖片字串: {str(e)}")
                        else:
                            # 處理非字符串類型的圖片數據
                            try:
                                if hasattr(pic_data, 'read'):
                                    question["picture"] = pic_data.read()
                                elif hasattr(pic_data, 'tobytes'):
                                    question["picture"] = pic_data.tobytes()
                                else:
                                    try:
                                        question["picture"] = bytes(pic_data)
                                    except:
                                        question["picture"] = None
                                
                                if question["picture"] is not None:
                                    logger.info(f"成功處理非字串圖片，大小: {len(question['picture'])} 位元組")
                            except Exception as e:
                                question["picture"] = None
                                logger.warning(f"無法處理圖片數據: {str(e)}")
                
                # 檢查圖片是否成功處理
                if question["picture"] is not None:
                    logger.info(f"題目 {question['question_no']} 成功添加圖片，大小: {len(question['picture'])} 位元組")
                else:
                    logger.info(f"題目 {question['question_no']} 沒有圖片")
                
                # 輸出調試信息，檢查每個題目的欄位
                logger.info(f"成功解析題目 {question['question_no']}: {question['question_text'][:30]}... 選項數: {len([x for x in [question['option_a'], question['option_b'], question['option_c'], question['option_d']] if x])}, 答案: {question['correct_answer'][:30]}...")
                    
                questions.append(question)
                valid_rows += 1
            
            # 處理結果統計
            logger.info(f"Excel 解析結果: 總計 {total_rows} 行, 有效 {valid_rows} 行, 跳過 {skipped_rows} 行")
            if skipped_rows > 0:
                for reason, count in skipped_reasons.items():
                    logger.info(f"- 跳過原因 '{reason}': {count} 行")
            
            return questions
            
        except Exception as e:
            logger.error(f"解析 Excel 檔案時發生錯誤: {str(e)}", exc_info=True)
            raise e
